import csv
import io
import json
import os
import shutil
from datetime import datetime

import openpyxl
from flask import Blueprint, request, jsonify, session, send_file
from database import get_db
import database
from routes.auth import login_required, manager_required

data_bp = Blueprint('data_management', __name__, url_prefix='/api')

ENTITIES = [
    {
        'id': 'products',
        'label': 'Products',
        'cols': ['name', 'sku', 'barcode', 'cost_price', 'base_price', 'category', 'commission_class', 'low_stock', 'stock', 'brand', 'make', 'color'],
        'required': ['name', 'sku'],
        'table': 'products',
        'import_endpoint': True,
    },
    {
        'id': 'customers',
        'label': 'Customers',
        'cols': ['name', 'phone', 'email', 'address', 'baby_name', 'baby_birth', 'notes', 'credit', 'credit_limit'],
        'required': ['name'],
        'table': 'customers',
        'import_endpoint': True,
    },
    {
        'id': 'suppliers',
        'label': 'Suppliers',
        'cols': ['name', 'phone', 'email', 'address', 'contact_person', 'notes', 'balance', 'company_phone', 'supplier_code', 'type', 'area', 'city', 'country', 'telephone', 'fax', 'account_no', 'due_days'],
        'required': ['name'],
        'table': 'suppliers',
        'import_endpoint': True,
    },
    {
        'id': 'sales',
        'label': 'Sales Invoices',
        'cols': ['receipt', 'subtotal', 'discount', 'discount_type', 'tax', 'total', 'payment', 'status', 'paid', 'customer_name', 'staff_name', 'notes', 'created_at', 'product_name', 'variant_label', 'sku', 'qty', 'price', 'item_total'],
        'required': ['receipt'],
        'table': 'sales',
        'import_endpoint': True,
    },
    {
        'id': 'purchases',
        'label': 'Purchase Invoices',
        'cols': ['invoice_no', 'issue_date', 'due_date', 'supplier_name', 'description', 'invoice_amount', 'balance_due', 'status', 'item', 'product_id', 'qty', 'unit_price', 'item_total'],
        'required': ['invoice_no'],
        'table': 'purchase_invoices',
        'import_endpoint': True,
    },
    {
        'id': 'accounts',
        'label': 'Accounts',
        'cols': ['name', 'type', 'balance'],
        'required': ['name'],
        'table': 'accounts',
        'import_endpoint': True,
    },
    {
        'id': 'payments',
        'label': 'Payments',
        'cols': ['receipt', 'method', 'amount', 'reference', 'created_at'],
        'required': ['receipt', 'method', 'amount'],
        'table': 'payments',
        'import_endpoint': True,
    },
    {
        'id': 'expenses',
        'label': 'Expenses',
        'cols': ['category', 'amount', 'note', 'created_at'],
        'required': ['category', 'amount'],
        'table': 'expenses',
        'import_endpoint': True,
    },
]


def parse_upload(file):
    filename = file.filename.lower() if file.filename else ''
    rows = []
    if filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        data_rows = list(ws.iter_rows(values_only=True))
        if not data_rows:
            return None, 'Empty file'
        headers = [str(h).strip().lower() if h else '' for h in data_rows[0]]
        for row in data_rows[1:]:
            d = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                d[h] = str(val).strip() if val is not None else ''
            rows.append(d)
    elif filename.endswith('.csv'):
        try:
            content = file.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            file.seek(0)
            content = file.read().decode('utf-8', errors='replace')
        reader = csv.DictReader(io.StringIO(content))
        if not reader.fieldnames:
            return None, 'Empty or invalid CSV'
        reader.fieldnames = [h.strip().lower() for h in reader.fieldnames]
        for row in reader:
            d = {}
            for k, v in row.items():
                d[k.strip().lower()] = v.strip() if v else ''
            rows.append(d)
    else:
        return None, 'Unsupported file format. Use .csv or .xlsx'
    return rows, None


@data_bp.route('/data/entities')
@login_required
def list_entities():
    return jsonify([{
        'id': e['id'],
        'label': e['label'],
        'cols': e['cols'],
        'required': e['required'],
        'count': _count_entity(e['id']),
    } for e in ENTITIES])


def _count_entity(eid):
    with get_db() as db:
        try:
            if eid == 'products':
                return db.execute('SELECT COUNT(*) FROM products').fetchone()[0]
            elif eid == 'customers':
                return db.execute('SELECT COUNT(*) FROM customers').fetchone()[0]
            elif eid == 'suppliers':
                return db.execute('SELECT COUNT(*) FROM suppliers').fetchone()[0]
            elif eid == 'sales':
                return db.execute('SELECT COUNT(*) FROM sales').fetchone()[0]
            elif eid == 'purchases':
                return db.execute('SELECT COUNT(*) FROM purchase_invoices').fetchone()[0]
            elif eid == 'accounts':
                return db.execute('SELECT COUNT(*) FROM accounts').fetchone()[0]
            elif eid == 'payments':
                return db.execute('SELECT COUNT(*) FROM payments').fetchone()[0]
            elif eid == 'expenses':
                return db.execute('SELECT COUNT(*) FROM expenses').fetchone()[0]
        except Exception:
            return 0
    return 0


@data_bp.route('/data/template/<entity>')
@login_required
def download_template(entity):
    ent = next((e for e in ENTITIES if e['id'] == entity), None)
    if not ent:
        return jsonify({'error': 'Unknown entity'}), 400
    output = io.StringIO()
    output.write(','.join(ent['cols']) + '\n')
    return output.getvalue(), 200, {
        'Content-Type': 'text/csv',
        'Content-Disposition': f'attachment; filename={entity}_template.csv'
    }


def _write_entity_rows(entity, write_row):
    with get_db() as db:
        if entity == 'products':
            rows = [dict(r) for r in db.execute(
                'SELECT p.*, COALESCE((SELECT SUM(v2.stock) FROM variants v2 WHERE v2.product_id=p.id), 0) as stock FROM products p ORDER BY p.id'
            ).fetchall()]
            for r in rows:
                write_row([r['name'], r['sku'], r['barcode'] or '', r['cost_price'], r['base_price'], r['category'] or '', r['commission_class'] or '', r['low_stock'], r['stock'], r.get('brand') or '', r.get('make') or '', r.get('color') or ''])
        elif entity == 'customers':
            rows = [dict(r) for r in db.execute('SELECT * FROM customers ORDER BY id').fetchall()]
            for r in rows:
                write_row([r['name'], r['phone'], r['email'], r['address'], r['baby_name'], r['baby_birth'], r['notes'], r['credit'], r.get('credit_limit') or ''])
        elif entity == 'suppliers':
            rows = [dict(r) for r in db.execute('SELECT * FROM suppliers ORDER BY id').fetchall()]
            for r in rows:
                write_row([r['name'], r['phone'], r['email'], r['address'], r['contact_person'], r['notes'], r['balance'], r.get('company_phone', ''),
                           r.get('supplier_code', ''), r.get('type', ''), r.get('area', ''), r.get('city', ''), r.get('country', ''),
                           r.get('telephone', ''), r.get('fax', ''), r.get('account_no', ''), r.get('due_days', 0)])
        elif entity == 'sales':
            rows = [dict(r) for r in db.execute(
                'SELECT s.*, si.product_name, si.variant_label, si.sku as item_sku, si.quantity, si.price, si.total as item_total '
                'FROM sales s JOIN sale_items si ON si.sale_id = s.id ORDER BY s.id, si.id'
            ).fetchall()]
            for r in rows:
                write_row([r['receipt'], r['subtotal'], r['discount'], r['discount_type'], r['tax'], r['total'], r['payment'], r['status'], r['paid'], r['customer_name'], r['staff_name'], r['notes'], r['created_at'], r['product_name'], r['variant_label'], r['item_sku'], r['quantity'], r['price'], r['item_total']])
        elif entity == 'purchases':
            rows = [dict(r) for r in db.execute(
                'SELECT pi.*, s.name as supplier_name, pii.item, pii.product_id, pii.qty, pii.unit_price, pii.total as item_total '
                'FROM purchase_invoices pi '
                'LEFT JOIN suppliers s ON s.id = pi.supplier_id '
                'JOIN purchase_invoice_items pii ON pii.invoice_id = pi.id ORDER BY pi.id, pii.line_number'
            ).fetchall()]
            for r in rows:
                write_row([r['invoice_no'], r['issue_date'], r.get('due_date', ''), r['supplier_name'] or '', r.get('description', ''), r['invoice_amount'], r['balance_due'], r['status'], r['item'] or '', str(r.get('product_id') or ''), r['qty'], r['unit_price'], r['item_total']])
        elif entity == 'accounts':
            rows = [dict(r) for r in db.execute('SELECT * FROM accounts ORDER BY id').fetchall()]
            for r in rows:
                write_row([r['name'], r['type'], r['balance']])
        elif entity == 'payments':
            rows = [dict(r) for r in db.execute(
                'SELECT p.*, s.receipt FROM payments p JOIN sales s ON s.id = p.sale_id ORDER BY p.id'
            ).fetchall()]
            for r in rows:
                write_row([r['receipt'], r['method'], r['amount'], r['reference'], r['created_at']])
        elif entity == 'expenses':
            rows = [dict(r) for r in db.execute('SELECT * FROM expenses ORDER BY id').fetchall()]
            for r in rows:
                write_row([r['category'], r['amount'], r['note'], r['created_at']])


@data_bp.route('/data/export/<entity>')
@login_required
def export_data(entity):
    ent = next((e for e in ENTITIES if e['id'] == entity), None)
    if not ent:
        return jsonify({'error': 'Unknown entity'}), 400
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(ent['cols'])
    _write_entity_rows(entity, lambda row: writer.writerow(row))
    return output.getvalue(), 200, {
        'Content-Type': 'text/csv',
        'Content-Disposition': f'attachment; filename={entity}_export.csv'
    }


@data_bp.route('/data/export-batch', methods=['POST'])
@login_required
def export_batch():
    d = request.get_json() or {}
    ids = d.get('entities', [])
    if not ids:
        return jsonify({'error': 'No entities selected'}), 400
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for eid in ids:
        ent = next((e for e in ENTITIES if e['id'] == eid), None)
        if not ent:
            continue
        ws = wb.create_sheet(title=ent['label'][:31])
        ws.append(ent['cols'])
        _write_entity_rows(eid, lambda row, ws=ws: ws.append(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='jitm_export.xlsx')


@data_bp.route('/data/import/<entity>', methods=['POST'])
@login_required
@manager_required
def import_data(entity):
    ent = next((e for e in ENTITIES if e['id'] == entity), None)
    if not ent:
        return jsonify({'error': 'Unknown entity'}), 400
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file uploaded'}), 400
    rows, err = parse_upload(file)
    if err:
        return jsonify({'error': err}), 400
    if not rows:
        return jsonify({'error': 'No data rows found'}), 400
    header_cols = set(rows[0].keys())
    missing = [c for c in ent['required'] if c not in header_cols]
    if missing:
        return jsonify({'error': f'Missing required columns: {", ".join(missing)}'}), 400
    created = 0
    updated = 0
    changes = []
    skipped = 0
    errors = []
    staff_name = session.get('name', '')
    with get_db() as db:
        db.execute('BEGIN IMMEDIATE')
        if entity == 'products':
            all_products = {r['sku']: dict(r) for r in db.execute('SELECT p.*, v.id as vid, v.stock as vstock FROM products p LEFT JOIN variants v ON v.product_id=p.id AND v.id=(SELECT MIN(v2.id) FROM variants v2 WHERE v2.product_id=p.id)').fetchall()}
            existing_skus = set(all_products.keys())
            existing_barcodes = set(r['barcode'] for r in db.execute('SELECT barcode FROM products WHERE barcode IS NOT NULL').fetchall())
            for idx, row in enumerate(rows, start=2):
                name = (row.get('name') or '').strip()
                sku = (row.get('sku') or '').strip()
                if not name or not sku:
                    errors.append(f'Row {idx}: name and sku are required')
                    continue
                try:
                    cost_price = float(row.get('cost_price', 0) or 0)
                    base_price = float(row.get('base_price', 0) or 0)
                    low_stock = int(row.get('low_stock', 5) or 5)
                    stock = int(row.get('stock', 0) or 0)
                except (ValueError, TypeError) as e:
                    errors.append(f'Row {idx}: invalid numeric value - {str(e)}')
                    continue
                barcode = (row.get('barcode') or '').strip() or None
                commission_class = (row.get('commission_class') or '').strip() or None
                category = (row.get('category') or '').strip()
                brand = (row.get('brand') or '').strip()
                make = (row.get('make') or '').strip()
                color = (row.get('color') or '').strip()

                if sku in existing_skus:
                    p = all_products[sku]
                    changed = []
                    updates = []
                    params = []
                    if p['name'] != name:
                        changed.append('name')
                        updates.append('name=?')
                        params.append(name)
                    if (p.get('category') or '') != category:
                        changed.append('category')
                        updates.append('category=?')
                        params.append(category)
                    if (p['base_price'] or 0) != base_price:
                        changed.append('base_price')
                        updates.append('base_price=?')
                        params.append(base_price)
                    if (p['cost_price'] or 0) != cost_price:
                        changed.append('cost_price')
                        updates.append('cost_price=?')
                        params.append(cost_price)
                    if (p.get('low_stock') or 5) != low_stock:
                        changed.append('low_stock')
                        updates.append('low_stock=?')
                        params.append(low_stock)
                    first_bc = p.get('barcode')
                    if (first_bc or '') != (barcode or ''):
                        if barcode and barcode in existing_barcodes and barcode != first_bc:
                            errors.append(f'Row {idx}: barcode "{barcode}" belongs to another product')
                            continue
                        changed.append('barcode')
                        updates.append('barcode=?')
                        params.append(barcode)
                    if (p.get('commission_class') or '') != (commission_class or ''):
                        changed.append('commission_class')
                        updates.append('commission_class=?')
                        params.append(commission_class)
                    if (p.get('brand') or '') != brand:
                        changed.append('brand')
                        updates.append('brand=?')
                        params.append(brand)
                    if (p.get('make') or '') != make:
                        changed.append('make')
                        updates.append('make=?')
                        params.append(make)
                    if (p.get('color') or '') != color:
                        changed.append('color')
                        updates.append('color=?')
                        params.append(color)
                    if updates:
                        params.append(p['id'])
                        db.execute(f'UPDATE products SET {",".join(updates)} WHERE id=?', params)
                    vid = p.get('vid')
                    old_stock = p.get('vstock') or 0
                    if vid and stock != old_stock:
                        db.execute('UPDATE variants SET stock=? WHERE id=?', (stock, vid))
                        changed.append('stock')
                        if stock > old_stock:
                            qty = stock - old_stock
                            db.execute(
                                'INSERT INTO restock_log (variant_id, old_stock, new_stock, qty_added, cost, note, staff_name) VALUES (?,?,?,?,?,?,?)',
                                (vid, old_stock, stock, qty, cost_price, 'Import stock update', staff_name)
                            )
                    if changed:
                        updated += 1
                        changes.append(f'Row {idx} ({sku}): {", ".join(changed)} updated')
                    else:
                        skipped += 1
                else:
                    if barcode and barcode in existing_barcodes:
                        errors.append(f'Row {idx}: barcode "{barcode}" already exists')
                        continue
                    try:
                        cur = db.execute(
                            'INSERT INTO products (name, category, base_price, cost_price, sku, barcode, low_stock, commission_class, brand, make, color) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                            (name, category, base_price, cost_price, sku, barcode, low_stock, commission_class, brand, make, color)
                        )
                        pid = cur.lastrowid
                        cur2 = db.execute('INSERT INTO variants (product_id, sku, stock) VALUES (?,?,?)', (pid, sku, stock))
                        if stock > 0:
                            db.execute('INSERT INTO restock_log (variant_id, old_stock, new_stock, qty_added, cost, note, staff_name) VALUES (?,?,?,?,?,?,?)',
                                       (cur2.lastrowid, 0, stock, stock, cost_price, 'Opening stock', staff_name))
                        existing_skus.add(sku)
                        all_products[sku] = {'id': pid, 'vid': cur2.lastrowid}
                        if barcode:
                            existing_barcodes.add(barcode)
                        created += 1
                    except Exception as e:
                        errors.append(f'Row {idx}: {str(e)}')
        elif entity == 'customers':
            existing_names = set(r['name'] for r in db.execute('SELECT name FROM customers').fetchall())
            for idx, row in enumerate(rows, start=2):
                name = (row.get('name') or '').strip()
                if not name:
                    errors.append(f'Row {idx}: name is required')
                    continue
                if name.lower() in {n.lower() for n in existing_names}:
                    skipped += 1
                    continue
                try:
                    credit = float(row.get('credit', 0) or 0)
                    credit_limit = float(row.get('credit_limit', 0) or 0) or None
                except (ValueError, TypeError):
                    errors.append(f'Row {idx}: invalid credit value')
                    continue
                try:
                    db.execute(
                        'INSERT INTO customers (name, phone, email, address, baby_name, baby_birth, notes, credit, credit_limit) VALUES (?,?,?,?,?,?,?,?,?)',
                        (name, row.get('phone', ''), row.get('email', ''), row.get('address', ''),
                         row.get('baby_name', ''), row.get('baby_birth', ''), row.get('notes', ''), credit, credit_limit)
                    )
                    existing_names.add(name)
                    created += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
        elif entity == 'suppliers':
            existing_names = set(r['name'] for r in db.execute('SELECT name FROM suppliers').fetchall())
            for idx, row in enumerate(rows, start=2):
                name = (row.get('name') or '').strip()
                if not name:
                    errors.append(f'Row {idx}: name is required')
                    continue
                if name.lower() in {n.lower() for n in existing_names}:
                    skipped += 1
                    continue
                try:
                    balance = float(row.get('balance', 0) or 0)
                except (ValueError, TypeError):
                    errors.append(f'Row {idx}: invalid balance')
                    continue
                try:
                    db.execute(
                        'INSERT INTO suppliers (name, phone, email, address, contact_person, notes, balance, company_phone, supplier_code, type, area, city, country, telephone, fax, account_no, due_days) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                        (name, row.get('phone', ''), row.get('email', ''), row.get('address', ''),
                         row.get('contact_person', ''), row.get('notes', ''), balance, row.get('company_phone', ''),
                         row.get('supplier_code', ''), row.get('type', ''), row.get('area', ''),
                         row.get('city', ''), row.get('country', ''), row.get('telephone', ''),
                         row.get('fax', ''), row.get('account_no', ''), int(row.get('due_days', 0) or 0))
                    )
                    existing_names.add(name)
                    created += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
        elif entity == 'sales':
            for idx, row in enumerate(rows, start=2):
                receipt = (row.get('receipt') or '').strip()
                if not receipt:
                    errors.append(f'Row {idx}: receipt is required')
                    continue
                existing = db.execute('SELECT id FROM sales WHERE receipt=?', (receipt,)).fetchone()
                if not existing:
                    try:
                        subtotal = float(row.get('subtotal', 0) or 0)
                        discount = float(row.get('discount', 0) or 0)
                        discount_type = (row.get('discount_type') or 'percent').strip()
                        tax = float(row.get('tax', 0) or 0)
                        total = float(row.get('total', 0) or 0)
                        payment = (row.get('payment') or 'cash').strip()
                        status = (row.get('status') or 'Paid').strip()
                        paid = float(row.get('paid', 0) or 0)
                        customer_name = (row.get('customer_name') or '').strip()
                        staff_name = (row.get('staff_name') or staff_name).strip()
                        notes = (row.get('notes') or '').strip()
                        created_at = (row.get('created_at') or '').strip()
                        cur = db.execute(
                            'INSERT INTO sales (receipt, subtotal, discount, discount_type, tax, total, payment, status, paid, customer_name, staff_name, notes, created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
                            (receipt, subtotal, discount, discount_type, tax, total, payment, status, paid, customer_name, staff_name, notes, created_at or None)
                        )
                        sale_id = cur.lastrowid
                        product_name = (row.get('product_name') or '').strip()
                        if product_name:
                            variant_label = (row.get('variant_label') or '').strip()
                            item_sku = (row.get('sku') or '').strip()
                            try:
                                qty = int(row.get('qty', 1) or 1)
                                price = float(row.get('price', 0) or 0)
                                item_total = float(row.get('item_total', 0) or 0)
                            except (ValueError, TypeError):
                                errors.append(f'Row {idx}: invalid item numeric value')
                                continue
                            db.execute(
                                'INSERT INTO sale_items (sale_id, product_id, product_name, variant_label, sku, quantity, price, total) VALUES (?,?,?,?,?,?,?,?)',
                                (sale_id, 0, product_name, variant_label, item_sku, qty, price, item_total)
                            )
                    except Exception as e:
                        errors.append(f'Row {idx}: {str(e)}')
                        continue
                else:
                    sale_id = existing['id']
                    product_name = (row.get('product_name') or '').strip()
                    if product_name:
                        variant_label = (row.get('variant_label') or '').strip()
                        item_sku = (row.get('sku') or '').strip()
                        try:
                            qty = int(row.get('qty', 1) or 1)
                            price = float(row.get('price', 0) or 0)
                            item_total = float(row.get('item_total', 0) or 0)
                        except (ValueError, TypeError):
                            errors.append(f'Row {idx}: invalid item numeric value')
                            continue
                        db.execute(
                            'INSERT INTO sale_items (sale_id, product_id, product_name, variant_label, sku, quantity, price, total) VALUES (?,?,?,?,?,?,?,?)',
                            (sale_id, 0, product_name, variant_label, item_sku, qty, price, item_total)
                        )
                created += 1
        elif entity == 'purchases':
            for idx, row in enumerate(rows, start=2):
                invoice_no = (row.get('invoice_no') or '').strip()
                if not invoice_no:
                    errors.append(f'Row {idx}: invoice_no is required')
                    continue
                existing = db.execute('SELECT id FROM purchase_invoices WHERE invoice_no=?', (invoice_no,)).fetchone()
                if not existing:
                    try:
                        invoice_amount = float(row.get('invoice_amount', 0) or 0)
                        balance_due = float(row.get('balance_due', 0) or 0)
                        status = (row.get('status') or 'Unpaid').strip()
                        issue_date = (row.get('issue_date') or '').strip()
                        due_date = (row.get('due_date') or '').strip()
                        supplier_name = (row.get('supplier_name') or '').strip()
                        description = (row.get('description') or '').strip()
                        supplier_id = None
                        if supplier_name:
                            sup = db.execute('SELECT id FROM suppliers WHERE name=?', (supplier_name,)).fetchone()
                            if sup:
                                supplier_id = sup['id']
                    except (ValueError, TypeError):
                        errors.append(f'Row {idx}: invalid numeric value')
                        continue
                    try:
                        cur = db.execute(
                            'INSERT INTO purchase_invoices (invoice_no, issue_date, due_date, supplier_id, description, invoice_amount, balance_due, status) VALUES (?,?,?,?,?,?,?,?)',
                            (invoice_no, issue_date, due_date, supplier_id, description, invoice_amount, balance_due, status)
                        )
                        inv_id = cur.lastrowid
                        item = (row.get('item') or '').strip()
                        if item:
                            try:
                                qty = float(row.get('qty', 1) or 1)
                                unit_price = float(row.get('unit_price', 0) or 0)
                                item_total = float(row.get('item_total', 0) or 0)
                            except (ValueError, TypeError):
                                errors.append(f'Row {idx}: invalid item numeric value')
                                continue
                            product_id = None
                            try:
                                product_id = int(row.get('product_id')) if row.get('product_id', '').strip() else None
                            except (ValueError, TypeError):
                                pass
                            db.execute(
                                'INSERT INTO purchase_invoice_items (invoice_id, line_number, item, product_id, qty, unit_price, total) VALUES (?,?,?,?,?,?,?)',
                                (inv_id, 1, item, product_id, qty, unit_price, item_total)
                            )
                    except Exception as e:
                        errors.append(f'Row {idx}: {str(e)}')
                        continue
                else:
                    inv_id = existing['id']
                    item = (row.get('item') or '').strip()
                    if item:
                        try:
                            qty = float(row.get('qty', 1) or 1)
                            unit_price = float(row.get('unit_price', 0) or 0)
                            item_total = float(row.get('item_total', 0) or 0)
                        except (ValueError, TypeError):
                            errors.append(f'Row {idx}: invalid item numeric value')
                            continue
                        max_ln = db.execute('SELECT COALESCE(MAX(line_number),0) as ln FROM purchase_invoice_items WHERE invoice_id=?', (inv_id,)).fetchone()['ln']
                        product_id = None
                        try:
                            product_id = int(row.get('product_id')) if row.get('product_id', '').strip() else None
                        except (ValueError, TypeError):
                            pass
                        db.execute(
                            'INSERT INTO purchase_invoice_items (invoice_id, line_number, item, product_id, qty, unit_price, total) VALUES (?,?,?,?,?,?,?)',
                            (inv_id, max_ln + 1, item, product_id, qty, unit_price, item_total)
                        )
                created += 1
        elif entity == 'accounts':
            existing_names = set(r['name'] for r in db.execute('SELECT name FROM accounts').fetchall())
            for idx, row in enumerate(rows, start=2):
                name = (row.get('name') or '').strip()
                if not name:
                    errors.append(f'Row {idx}: name is required')
                    continue
                if name.lower() in {n.lower() for n in existing_names}:
                    skipped += 1
                    continue
                try:
                    balance = float(row.get('balance', 0) or 0)
                except (ValueError, TypeError):
                    errors.append(f'Row {idx}: invalid balance')
                    continue
                acct_type = (row.get('type') or 'cash').strip()
                try:
                    db.execute('INSERT INTO accounts (name, type, balance) VALUES (?,?,?)', (name, acct_type, balance))
                    existing_names.add(name)
                    created += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
        elif entity == 'payments':
            for idx, row in enumerate(rows, start=2):
                receipt = (row.get('receipt') or '').strip()
                method = (row.get('method') or '').strip()
                try:
                    amount = float(row.get('amount', 0) or 0)
                except (ValueError, TypeError):
                    errors.append(f'Row {idx}: invalid amount')
                    continue
                if not receipt or not method or amount <= 0:
                    errors.append(f'Row {idx}: receipt, method, and amount>0 are required')
                    continue
                sale = db.execute('SELECT id FROM sales WHERE receipt=?', (receipt,)).fetchone()
                if not sale:
                    errors.append(f'Row {idx}: sale with receipt "{receipt}" not found')
                    continue
                try:
                    db.execute(
                        'INSERT INTO payments (sale_id, method, amount, reference, created_at) VALUES (?,?,?,?,?)',
                        (sale['id'], method, amount, row.get('reference', ''), (row.get('created_at') or '').strip() or None)
                    )
                    created += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
        elif entity == 'expenses':
            for idx, row in enumerate(rows, start=2):
                category = (row.get('category') or '').strip()
                try:
                    amount = float(row.get('amount', 0) or 0)
                except (ValueError, TypeError):
                    errors.append(f'Row {idx}: invalid amount')
                    continue
                if not category or amount <= 0:
                    errors.append(f'Row {idx}: category and amount>0 are required')
                    continue
                try:
                    db.execute(
                        'INSERT INTO expenses (category, amount, note, created_at) VALUES (?,?,?,?)',
                        (category, amount, (row.get('note') or '').strip(), (row.get('created_at') or '').strip() or None)
                    )
                    created += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
    return jsonify({'ok': True, 'created': created, 'updated': updated, 'changes': changes, 'skipped': skipped, 'errors': errors})


@data_bp.route('/data/db-path')
@login_required
def get_db_path():
    return jsonify({
        'current_path': database.DB,
        'config_file': database.CONFIG_FILE,
        'exists': os.path.exists(database.DB),
        'size': os.path.getsize(database.DB) if os.path.exists(database.DB) else 0,
    })


@data_bp.route('/data/db-path', methods=['PUT'])
@login_required
@manager_required
def change_db_path():
    d = request.get_json() or {}
    new_path = (d.get('path') or '').strip()
    if not new_path:
        return jsonify({'error': 'Path is required'}), 400
    new_path = os.path.abspath(new_path)
    if not new_path.endswith('.db'):
        new_path = os.path.join(new_path, 'jitm.db')
    dirname = os.path.dirname(new_path)
    if not os.path.exists(dirname):
        os.makedirs(dirname, exist_ok=True)
    old_path = database.DB
    if os.path.normpath(old_path) == os.path.normpath(new_path):
        return jsonify({'error': 'New path is the same as current path'}), 400
    if os.path.exists(old_path):
        for f in [old_path, old_path + '-wal', old_path + '-shm']:
            if os.path.exists(f):
                shutil.copy2(f, os.path.join(dirname, os.path.basename(f)))
    database.set_db_path(new_path)
    return jsonify({'ok': True, 'new_path': database.DB, 'message': 'Database path updated. Restart the app for the change to take effect.'})


_CLEAR_ORDER = {
    'products': ['restock_log', 'variants', 'products'],
    'customers': ['customers'],
    'suppliers': ['suppliers'],
    'sales': ['sale_items', 'payments', 'sales'],
    'purchases': ['purchase_invoice_items', 'purchase_invoices'],
    'accounts': ['account_transfers', 'transactions', 'accounts'],
    'payments': ['payments'],
    'expenses': ['expenses'],
}

ALL_TABLES = [
    'commission_classes', 'categories', 'sizes', 'users', 'suppliers',
    'accounts', 'customers', 'products', 'variants', 'employees',
    'purchase_invoices', 'purchase_invoice_items', 'purchase_returns',
    'purchase_return_items', 'sales', 'sale_items', 'payments',
    'restock_log', 'account_transfers', 'transactions', 'expenses',
    'settings', 'attendance', 'raw_materials', 'bom', 'recipe_profiles',
    'recipe_profile_items', 'production_orders', 'production_order_items',
    'material_adjustments', 'material_transfers',
]

_CLEAR_ALL_ORDER = [
    'attendance',
    'restock_log',
    'payments',
    'sale_items',
    'sales',
    'purchase_return_items',
    'purchase_returns',
    'purchase_invoice_items',
    'purchase_invoices',
    'account_transfers',
    'transactions',
    'production_order_items',
    'production_orders',
    'bom',
    'recipe_profile_items',
    'recipe_profiles',
    'material_transfers',
    'material_adjustments',
    'raw_materials',
    'variants',
    'products',
    'employees',
    'customers',
    'accounts',
    'suppliers',
    'users',
    'sizes',
    'categories',
    'commission_classes',
    'expenses',
    'settings',
]

_INSERT_ORDER = list(reversed(_CLEAR_ALL_ORDER))


def _clear_all():
    with get_db() as db:
        db.execute('PRAGMA foreign_keys=OFF')
        try:
            for t in _CLEAR_ALL_ORDER:
                db.execute(f'DELETE FROM {t}')
        finally:
            db.execute('PRAGMA foreign_keys=ON')


@data_bp.route('/data/backup')
@login_required
@manager_required
def backup_data():
    backup = {
        'version': 2,
        'created_at': datetime.now().isoformat(),
        'tables': {},
    }
    with get_db() as db:
        for table in ALL_TABLES:
            try:
                rows = [dict(r) for r in db.execute(f'SELECT * FROM {table} ORDER BY id').fetchall()]
                backup['tables'][table] = rows
            except Exception:
                backup['tables'][table] = []
    return send_file(
        io.BytesIO(json.dumps(backup, indent=2).encode('utf-8')),
        mimetype='application/json',
        as_attachment=True,
        download_name='jitm_backup.json'
    )


@data_bp.route('/data/restore', methods=['POST'])
@login_required
@manager_required
def restore_data():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file uploaded'}), 400
    try:
        data = json.loads(file.read().decode('utf-8'))
    except Exception as e:
        return jsonify({'error': f'Invalid JSON: {str(e)}'}), 400
    if not isinstance(data, dict) or 'tables' not in data:
        return jsonify({'error': 'Invalid backup format: missing "tables" key'}), 400
    tables = data.get('tables', {})
    counts = {}
    with get_db() as db:
        db.execute('PRAGMA foreign_keys=OFF')
        try:
            for t in _CLEAR_ALL_ORDER:
                db.execute(f'DELETE FROM {t}')
            for table in _INSERT_ORDER:
                rows = tables.get(table, [])
                if not rows:
                    counts[table] = 0
                    continue
                if table == 'settings':
                    for row in rows:
                        db.execute(
                            'INSERT OR REPLACE INTO settings (key, value) VALUES (?,?)',
                            (row['key'], row['value'])
                        )
                    counts[table] = len(rows)
                    continue
                col_names = list(rows[0].keys()) if rows else []
                if not col_names:
                    counts[table] = 0
                    continue
                placeholders = ','.join(['?' for _ in col_names])
                cols_csv = ','.join(col_names)
                for row in rows:
                    values = [row.get(c) for c in col_names]
                    db.execute(f'INSERT INTO {table} ({cols_csv}) VALUES ({placeholders})', values)
                counts[table] = len(rows)
        except Exception as e:
            return jsonify({'error': f'Restore failed at table "{table}": {str(e)}'}), 500
        finally:
            db.execute('PRAGMA foreign_keys=ON')
    return jsonify({'ok': True, 'counts': counts, 'message': 'Database restored successfully'})


def _clear_entity(entity_id):
    with get_db() as db:
        db.execute('PRAGMA foreign_keys=OFF')
        try:
            tables = _CLEAR_ORDER.get(entity_id, [entity_id])
            for t in tables:
                db.execute(f'DELETE FROM {t}')
        finally:
            db.execute('PRAGMA foreign_keys=ON')


@data_bp.route('/data/clear/<entity>', methods=['DELETE'])
@login_required
@manager_required
def clear_entity(entity):
    ent = next((e for e in ENTITIES if e['id'] == entity), None)
    if not ent:
        return jsonify({'error': 'Unknown entity'}), 400
    _clear_entity(entity)
    return jsonify({'ok': True, 'message': f'All {ent["label"]} records cleared.'})


@data_bp.route('/data/import-replace/<entity>', methods=['POST'])
@login_required
@manager_required
def import_replace(entity):
    ent = next((e for e in ENTITIES if e['id'] == entity), None)
    if not ent:
        return jsonify({'error': 'Unknown entity'}), 400
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file uploaded'}), 400
    rows, err = parse_upload(file)
    if err:
        return jsonify({'error': err}), 400
    if not rows:
        return jsonify({'error': 'No data rows found'}), 400
    header_cols = set(rows[0].keys())
    missing = [c for c in ent['required'] if c not in header_cols]
    if missing:
        return jsonify({'error': f'Missing required columns: {", ".join(missing)}'}), 400

    _clear_entity(entity)

    created = 0
    updated = 0
    changes = []
    skipped = 0
    errors = []
    staff_name = session.get('name', '')
    with get_db() as db:
        db.execute('PRAGMA foreign_keys=OFF')
        try:
            db.execute('BEGIN IMMEDIATE')
            if entity == 'products':
                all_products = {}
                existing_barcodes = set()
                for idx, row in enumerate(rows, start=2):
                    name = (row.get('name') or '').strip()
                    sku = (row.get('sku') or '').strip()
                    if not name or not sku:
                        errors.append(f'Row {idx}: name and sku are required')
                        continue
                    try:
                        cost_price = float(row.get('cost_price', 0) or 0)
                        base_price = float(row.get('base_price', 0) or 0)
                        low_stock = int(row.get('low_stock', 5) or 5)
                        stock = int(row.get('stock', 0) or 0)
                    except (ValueError, TypeError) as e:
                        errors.append(f'Row {idx}: invalid numeric value - {str(e)}')
                        continue
                    barcode = (row.get('barcode') or '').strip() or None
                    commission_class = (row.get('commission_class') or '').strip() or None
                    category = (row.get('category') or '').strip()
                    brand = (row.get('brand') or '').strip()
                    make = (row.get('make') or '').strip()
                    color = (row.get('color') or '').strip()

                    if sku in all_products:
                        p = all_products[sku]
                        changed = []
                        updates = []
                        params = []
                        if p['name'] != name:
                            changed.append('name')
                            updates.append('name=?')
                            params.append(name)
                        if (p.get('category') or '') != category:
                            changed.append('category')
                            updates.append('category=?')
                            params.append(category)
                        if (p['base_price'] or 0) != base_price:
                            changed.append('base_price')
                            updates.append('base_price=?')
                            params.append(base_price)
                        if (p['cost_price'] or 0) != cost_price:
                            changed.append('cost_price')
                            updates.append('cost_price=?')
                            params.append(cost_price)
                        if (p.get('low_stock') or 5) != low_stock:
                            changed.append('low_stock')
                            updates.append('low_stock=?')
                            params.append(low_stock)
                        if (p.get('commission_class') or '') != (commission_class or ''):
                            changed.append('commission_class')
                            updates.append('commission_class=?')
                            params.append(commission_class)
                        if (p.get('brand') or '') != brand:
                            changed.append('brand')
                            updates.append('brand=?')
                            params.append(brand)
                        if (p.get('make') or '') != make:
                            changed.append('make')
                            updates.append('make=?')
                            params.append(make)
                        if (p.get('color') or '') != color:
                            changed.append('color')
                            updates.append('color=?')
                            params.append(color)
                        if barcode and barcode in existing_barcodes and barcode != p.get('barcode'):
                            errors.append(f'Row {idx}: barcode "{barcode}" belongs to another product')
                            continue
                        if (p.get('barcode') or '') != (barcode or ''):
                            changed.append('barcode')
                            updates.append('barcode=?')
                            params.append(barcode or None)
                        if updates:
                            params.append(p['id'])
                            db.execute(f'UPDATE products SET {",".join(updates)} WHERE id=?', params)
                        vid = p.get('vid')
                        old_stock = p.get('vstock') or 0
                        if vid and stock != old_stock:
                            db.execute('UPDATE variants SET stock=? WHERE id=?', (stock, vid))
                            changed.append('stock')
                            if stock > old_stock:
                                qty = stock - old_stock
                                db.execute(
                                    'INSERT INTO restock_log (variant_id, old_stock, new_stock, qty_added, cost, note, staff_name) VALUES (?,?,?,?,?,?,?)',
                                    (vid, old_stock, stock, qty, cost_price, 'Import stock update', staff_name)
                                )
                        if changed:
                            updated += 1
                            changes.append(f'Row {idx} ({sku}): {", ".join(changed)} updated')
                        else:
                            skipped += 1
                    else:
                        if barcode and barcode in existing_barcodes:
                            errors.append(f'Row {idx}: barcode "{barcode}" already exists')
                            continue
                        try:
                            cur = db.execute(
                                'INSERT INTO products (name, category, base_price, cost_price, sku, barcode, low_stock, commission_class, brand, make, color) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                                (name, category, base_price, cost_price, sku, barcode, low_stock, commission_class, brand, make, color)
                            )
                            pid = cur.lastrowid
                            cur2 = db.execute('INSERT INTO variants (product_id, sku, stock) VALUES (?,?,?)', (pid, sku, stock))
                            if stock > 0:
                                db.execute('INSERT INTO restock_log (variant_id, old_stock, new_stock, qty_added, cost, note, staff_name) VALUES (?,?,?,?,?,?,?)',
                                           (cur2.lastrowid, 0, stock, stock, cost_price, 'Opening stock', staff_name))
                            all_products[sku] = {'id': pid, 'vid': cur2.lastrowid, 'name': name,
                                'base_price': base_price, 'cost_price': cost_price, 'low_stock': low_stock,
                                'vstock': stock, 'barcode': barcode, 'category': category,
                                'commission_class': commission_class, 'brand': brand, 'make': make, 'color': color}
                            if barcode:
                                existing_barcodes.add(barcode)
                            created += 1
                        except Exception as e:
                            errors.append(f'Row {idx}: {str(e)}')
            elif entity == 'customers':
                for idx, row in enumerate(rows, start=2):
                    name = (row.get('name') or '').strip()
                    if not name:
                        errors.append(f'Row {idx}: name is required')
                        continue
                    try:
                        credit = float(row.get('credit', 0) or 0)
                        credit_limit = float(row.get('credit_limit', 0) or 0) or None
                    except (ValueError, TypeError):
                        errors.append(f'Row {idx}: invalid credit value')
                        continue
                    try:
                        db.execute(
                            'INSERT INTO customers (name, phone, email, address, baby_name, baby_birth, notes, credit, credit_limit) VALUES (?,?,?,?,?,?,?,?,?)',
                            (name, row.get('phone', ''), row.get('email', ''), row.get('address', ''),
                             row.get('baby_name', ''), row.get('baby_birth', ''), row.get('notes', ''), credit, credit_limit)
                        )
                        created += 1
                    except Exception as e:
                        errors.append(f'Row {idx}: {str(e)}')
            elif entity == 'suppliers':
                for idx, row in enumerate(rows, start=2):
                    name = (row.get('name') or '').strip()
                    if not name:
                        errors.append(f'Row {idx}: name is required')
                        continue
                    try:
                        balance = float(row.get('balance', 0) or 0)
                    except (ValueError, TypeError):
                        errors.append(f'Row {idx}: invalid balance')
                        continue
                    try:
                        db.execute(
                            'INSERT INTO suppliers (name, phone, email, address, contact_person, notes, balance, company_phone, supplier_code, type, area, city, country, telephone, fax, account_no, due_days) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                            (name, row.get('phone', ''), row.get('email', ''), row.get('address', ''),
                             row.get('contact_person', ''), row.get('notes', ''), balance, row.get('company_phone', ''),
                             row.get('supplier_code', ''), row.get('type', ''), row.get('area', ''),
                             row.get('city', ''), row.get('country', ''), row.get('telephone', ''),
                             row.get('fax', ''), row.get('account_no', ''), int(row.get('due_days', 0) or 0))
                        )
                        created += 1
                    except Exception as e:
                        errors.append(f'Row {idx}: {str(e)}')
            elif entity == 'sales':
                for idx, row in enumerate(rows, start=2):
                    receipt = (row.get('receipt') or '').strip()
                    if not receipt:
                        errors.append(f'Row {idx}: receipt is required')
                        continue
                    try:
                        subtotal = float(row.get('subtotal', 0) or 0)
                        discount = float(row.get('discount', 0) or 0)
                        discount_type = (row.get('discount_type') or 'percent').strip()
                        tax = float(row.get('tax', 0) or 0)
                        total = float(row.get('total', 0) or 0)
                        payment = (row.get('payment') or 'cash').strip()
                        status = (row.get('status') or 'Paid').strip()
                        paid = float(row.get('paid', 0) or 0)
                        customer_name = (row.get('customer_name') or '').strip()
                        row_staff = (row.get('staff_name') or staff_name).strip()
                        notes = (row.get('notes') or '').strip()
                        created_at = (row.get('created_at') or '').strip()
                        cur = db.execute(
                            'INSERT INTO sales (receipt, subtotal, discount, discount_type, tax, total, payment, status, paid, customer_name, staff_name, notes, created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
                            (receipt, subtotal, discount, discount_type, tax, total, payment, status, paid, customer_name, row_staff, notes, created_at or None)
                        )
                        sale_id = cur.lastrowid
                        product_name = (row.get('product_name') or '').strip()
                        if product_name:
                            variant_label = (row.get('variant_label') or '').strip()
                            item_sku = (row.get('sku') or '').strip()
                            qty = int(row.get('qty', 1) or 1)
                            price = float(row.get('price', 0) or 0)
                            item_total = float(row.get('item_total', 0) or 0)
                            db.execute(
                                'INSERT INTO sale_items (sale_id, product_id, product_name, variant_label, sku, quantity, price, total) VALUES (?,?,?,?,?,?,?,?)',
                                (sale_id, 0, product_name, variant_label, item_sku, qty, price, item_total)
                            )
                        created += 1
                    except Exception as e:
                        errors.append(f'Row {idx}: {str(e)}')
            elif entity == 'purchases':
                for idx, row in enumerate(rows, start=2):
                    invoice_no = (row.get('invoice_no') or '').strip()
                    if not invoice_no:
                        errors.append(f'Row {idx}: invoice_no is required')
                        continue
                    try:
                        invoice_amount = float(row.get('invoice_amount', 0) or 0)
                        balance_due = float(row.get('balance_due', 0) or 0)
                        status = (row.get('status') or 'Unpaid').strip()
                        issue_date = (row.get('issue_date') or '').strip()
                        due_date = (row.get('due_date') or '').strip()
                        supplier_name = (row.get('supplier_name') or '').strip()
                        description = (row.get('description') or '').strip()
                        supplier_id = None
                        if supplier_name:
                            sup = db.execute('SELECT id FROM suppliers WHERE name=?', (supplier_name,)).fetchone()
                            if sup:
                                supplier_id = sup['id']
                        cur = db.execute(
                            'INSERT INTO purchase_invoices (invoice_no, issue_date, due_date, supplier_id, description, invoice_amount, balance_due, status) VALUES (?,?,?,?,?,?,?,?)',
                            (invoice_no, issue_date, due_date, supplier_id, description, invoice_amount, balance_due, status)
                        )
                        inv_id = cur.lastrowid
                        item = (row.get('item') or '').strip()
                        if item:
                            qty = float(row.get('qty', 1) or 1)
                            unit_price = float(row.get('unit_price', 0) or 0)
                            item_total = float(row.get('item_total', 0) or 0)
                            product_id = None
                            try:
                                product_id = int(row.get('product_id')) if row.get('product_id', '').strip() else None
                            except (ValueError, TypeError):
                                pass
                            db.execute(
                                'INSERT INTO purchase_invoice_items (invoice_id, line_number, item, product_id, qty, unit_price, total) VALUES (?,?,?,?,?,?,?)',
                                (inv_id, 1, item, product_id, qty, unit_price, item_total)
                            )
                        created += 1
                    except Exception as e:
                        errors.append(f'Row {idx}: {str(e)}')
            elif entity == 'accounts':
                for idx, row in enumerate(rows, start=2):
                    name = (row.get('name') or '').strip()
                    if not name:
                        errors.append(f'Row {idx}: name is required')
                        continue
                    balance = float(row.get('balance', 0) or 0)
                    acct_type = (row.get('type') or 'cash').strip()
                    try:
                        db.execute('INSERT INTO accounts (name, type, balance) VALUES (?,?,?)', (name, acct_type, balance))
                        created += 1
                    except Exception as e:
                        errors.append(f'Row {idx}: {str(e)}')
            elif entity == 'payments':
                for idx, row in enumerate(rows, start=2):
                    receipt = (row.get('receipt') or '').strip()
                    method = (row.get('method') or '').strip()
                    amount = float(row.get('amount', 0) or 0)
                    if not receipt or not method or amount <= 0:
                        errors.append(f'Row {idx}: receipt, method, and amount>0 are required')
                        continue
                    sale = db.execute('SELECT id FROM sales WHERE receipt=?', (receipt,)).fetchone()
                    if not sale:
                        errors.append(f'Row {idx}: sale with receipt "{receipt}" not found')
                        continue
                    try:
                        db.execute(
                            'INSERT INTO payments (sale_id, method, amount, reference, created_at) VALUES (?,?,?,?,?)',
                            (sale['id'], method, amount, row.get('reference', ''), (row.get('created_at') or '').strip() or None)
                        )
                        created += 1
                    except Exception as e:
                        errors.append(f'Row {idx}: {str(e)}')
            elif entity == 'expenses':
                for idx, row in enumerate(rows, start=2):
                    category = (row.get('category') or '').strip()
                    amount = float(row.get('amount', 0) or 0)
                    if not category or amount <= 0:
                        errors.append(f'Row {idx}: category and amount>0 are required')
                        continue
                    try:
                        db.execute(
                            'INSERT INTO expenses (category, amount, note, created_at) VALUES (?,?,?,?)',
                            (category, amount, (row.get('note') or '').strip(), (row.get('created_at') or '').strip() or None)
                        )
                        created += 1
                    except Exception as e:
                        errors.append(f'Row {idx}: {str(e)}')
        finally:
            db.execute('PRAGMA foreign_keys=ON')
    return jsonify({'ok': True, 'created': created, 'updated': updated, 'changes': changes, 'skipped': skipped, 'errors': errors})
